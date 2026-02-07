#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extract instrumentation data from Excel file and add to database.
"""

import sqlite3
from pathlib import Path
from openpyxl import load_workbook

BASE_DIR = Path(__file__).parent.parent
DATA_DIR = BASE_DIR / "data"
DB_PATH = DATA_DIR / "archive.db"
EXCEL_PATH = BASE_DIR / "archive" / "files" / "Werke" / "2026-01-06  Liste kompositorisches Werk - Endfassung.xlsx"

def extract_instrumentation():
    """Extract instrumentation data from Excel column E."""
    print("[*] Loading Excel file...")
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    
    # Column E is the 5th column (index 5)
    # Start from row 4 (row 3 is header "Besetzung")
    instrumentations = set()
    work_instrumentations = {}
    
    print("[*] Extracting instrumentation data...")
    for row in range(4, ws.max_row + 1):
        # Column A: Work number
        work_number_cell = ws.cell(row=row, column=1)
        # Column E: Besetzung (Instrumentation)
        instrumentation_cell = ws.cell(row=row, column=5)
        
        work_number = work_number_cell.value
        instrumentation = instrumentation_cell.value
        
        if work_number and instrumentation:
            work_number_str = str(work_number).strip()
            instrumentation_str = str(instrumentation).strip()
            
            instrumentations.add(instrumentation_str)
            work_instrumentations[work_number_str] = instrumentation_str
    
    print(f"\n[*] Found {len(instrumentations)} unique instrumentations:")
    for inst in sorted(instrumentations):
        print(f"   - {inst}")
    
    print(f"\n[*] Found instrumentation data for {len(work_instrumentations)} works")
    
    return work_instrumentations, instrumentations

def update_database(work_instrumentations):
    """Add instrumentation column to database and populate it."""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    
    # Check if instrumentation column exists
    c.execute("PRAGMA table_info(works)")
    columns = [col[1] for col in c.fetchall()]
    
    if 'instrumentation' not in columns:
        print("[*] Adding instrumentation column to works table...")
        c.execute("ALTER TABLE works ADD COLUMN instrumentation TEXT")
        conn.commit()
    else:
        print("[*] Instrumentation column already exists")
    
    # Update works with instrumentation data
    print("[*] Updating works with instrumentation data...")
    updated = 0
    not_found = []
    
    for work_number, instrumentation in work_instrumentations.items():
        c.execute("SELECT id FROM works WHERE work_number = ?", (work_number,))
        result = c.fetchone()
        
        if result:
            work_id = result[0]
            c.execute("UPDATE works SET instrumentation = ? WHERE id = ?", 
                     (instrumentation, work_id))
            updated += 1
        else:
            not_found.append(work_number)
    
    conn.commit()
    
    print(f"[+] Updated {updated} works with instrumentation data")
    
    if not_found:
        print(f"\n[!] Could not find {len(not_found)} works in database:")
        for wn in not_found[:10]:
            print(f"   - Work {wn}")
        if len(not_found) > 10:
            print(f"   ... and {len(not_found) - 10} more")
    
    # Show statistics
    c.execute("SELECT COUNT(*) FROM works WHERE instrumentation IS NOT NULL")
    count_with_inst = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM works")
    total_works = c.fetchone()[0]
    
    print(f"\n[*] Statistics:")
    print(f"   Works with instrumentation: {count_with_inst}/{total_works}")
    
    conn.close()

def main():
    print("=" * 70)
    print("Emanuel Vogt Archive - Extract Instrumentation Data")
    print("=" * 70)
    print()
    
    if not EXCEL_PATH.exists():
        print(f"[-] Excel file not found: {EXCEL_PATH}")
        return
    
    try:
        work_instrumentations, instrumentations = extract_instrumentation()
        update_database(work_instrumentations)
        
        print("\n" + "=" * 70)
        print("[+] OPERATION COMPLETE")
        print("=" * 70)
        
        print("\nUnique instrumentations found:")
        for inst in sorted(instrumentations):
            print(f"  - {inst}")
            
    except Exception as e:
        print(f"\n[-] Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
